# 验证码识别
# 随机数
import json
import os.path
import pickle
import random
import string
import time

import openpyxl
import yaml
from PIL import Image
from selenium.webdriver.common.by import By

from lib.ShowapiRequest import ShowapiRequest

# 选择项目名下的目录
def select_dir(directory):
    pro_name = 'pom_mumu'
    # 本文件路径
    dirname = os.path.dirname(__file__)
    # 项目名称的下标
    pro_name_index = dirname.index(pro_name)
    # print(dirname)
    # print(pro_name_index)
    # 截取项目名称前的路径(换电脑了，可以获取新的路径进行拼接) 再加上 项目名 + 要存放的文件夹
    path_os = dirname[: pro_name_index + len(pro_name)] + os.sep + directory
    return path_os

# 日志
def get_logger():
    # 导入logging模块，用于日志记录
    import logging
    # 导入logging.handlers模块，它提供了许多处理器类，如TimedRotatingFileHandler
    import logging.handlers
    # 导入datetime模块，用于设置时间相关的参数
    import datetime


    # 获取一个名为'mylogger'的logger对象。如果此logger对象已存在，则直接返回它；否则，创建一个新的logger对象
    logger = logging.getLogger('mylogger')
    # 设置logger的日志级别为DEBUG，这意味着DEBUG级别及以上的日志都会被记录
    logger.setLevel(logging.DEBUG)


    # # 保存的路径
    path_os = select_dir('logs')
    # # 项目名称
    # pro_name = 'pom_mumu'
    # # 本文件路径
    # dirname = os.path.dirname(__file__)
    # # 项目名称的下标
    # pro_name_index = dirname.index(pro_name)
    # # print(dirname)
    # # print(pro_name_index)
    # # 截取项目名称前的路径(换电脑了，可以获取新的路径进行拼接) 再加上 项目名 + 要存放的文件夹
    # path_os = dirname[: pro_name_index + len(pro_name)] + os.sep + 'logs'
    all_log_path = path_os + os.sep + 'all.log'
    error_log_path = path_os + os.sep + 'error.log'


    # 创建一个TimedRotatingFileHandler对象，用于按时间间隔（每天午夜）将日志记录到'all.log'文件
    # 参数解释：
    #   - 'all.log': 日志文件名
    #   - when='midnight': 指定时间间隔为每天午夜
    #   - interval=1: 与when参数结合使用，这里表示每天
    #   - backupCount=7: 保留的备份日志文件数量，这里是7个
    #   - atTime=datetime.time(0, 0, 0, 0): 指定具体的时间（这里指定为午夜）
    rf_handler = logging.handlers.TimedRotatingFileHandler(all_log_path, when='midnight', interval=1, backupCount=7,
                                                           atTime=datetime.time(0, 0, 0, 0))
    # 设置日志格式
    rf_handler.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(message)s"))


    # 创建一个FileHandler对象，用于将ERROR级别及以上的日志记录到'error.log'文件
    f_handler = logging.FileHandler(error_log_path)
    # 设置FileHandler的日志级别为ERROR
    f_handler.setLevel(logging.ERROR)
    # 设置日志格式，这里还包含了文件名和行号信息
    f_handler.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(filename)s[:%(lineno)d] - %(message)s"))


    # 将两个处理器添加到logger对象中
    logger.addHandler(rf_handler)
    logger.addHandler(f_handler)
    # 返回配置好的logger对象
    return logger

# 使用第三方验证码接口
def get_code(driver, id):
    # 获取当前时间戳，用于生成验证码图片的文件名
    t = time.time()

    # 拼接截图保存的路径，这里假设在脚本的上级目录的screenshots文件夹下
    path = os.path.dirname(os.path.dirname(__file__)) + '\\screenshots'
    picture_name1 = path + '\\' + str(t) + '.png'

    # 对整个页面进行截图
    driver.save_screenshot(picture_name1)

    # 通过id定位验证码元素
    ce = driver.find_element(By.ID ,id)

    # 获取验证码元素的左上角x坐标
    left = ce.location['x']
    # 获取验证码元素的左上角y坐标
    top = ce.location['y']

    # 这里计算验证码元素的右下角x坐标的方式是错误的，应该是left + width
    right = ce.size['width'] + left  # 应该是 left + ce.size['width']
    # 这里计算验证码元素的下边缘y坐标的方式也是错误的，应该是top + height
    height = ce.size['height'] + top  # 应该是 top + ce.size['height']

    # 获取设备的像素比，用于后续计算截图区域
    dpr = driver.execute_script('return window.devicePixelRatio')

    # 打印设备的像素比
    print(dpr)

    # 打开整个页面的截图
    im = Image.open(picture_name1)

    # 根据验证码元素的坐标和设备的像素比进行截图
    # 但这里计算的区域是不准确的，因为我们之前的计算存在错误
    img = im.crop((left * dpr, top * dpr, right * dpr, height * dpr))

    # 获取新的时间戳，用于生成裁剪后的验证码图片的文件名
    t = time.time()

    # 拼接裁剪后的验证码图片保存的路径
    picture_name2 = path + '\\' + str(t) + '.png'

    # 保存裁剪后的验证码图片
    img.save(picture_name2)  # 这里就是截取到的验证码图片

    # 创建一个请求对象，用于向第三方验证码识别接口发送请求
    r = ShowapiRequest("http://route.showapi.com/184-4", "290728", "1bd001f23c874581aac4db788a92c71d")

    # 添加文件参数，即裁剪后的验证码图片
    r.addFilePara("image", picture_name2)

    # 添加请求体参数，指定验证码类型等
    r.addBodyPara("typeId", "34")
    r.addBodyPara("convert_to_jpg", "0")
    r.addBodyPara("needMorePrecise", "0")

    # 发送POST请求到第三方验证码识别接口
    res = r.post()

    # 解析返回的JSON数据，获取验证码识别结果
    text = res.json()['showapi_res_body']
    code = text['Result']

    # 返回识别到的验证码
    return code

# 生成随机字符串(8位)
def gen_random_str():
    rand_str = ''.join(random.sample(string.ascii_letters + string.digits, 8))
    return rand_str

# 保存cookie
def save_cookie(driver, path):
    with open(path, 'wb') as filehandler:
        cookies = driver.get_cookies()
        print(cookies)
        pickle.dump(cookies, filehandler)

# 加载cookie
def load_cookie(driver, path):
    with open(path, 'rb') as cookiesfile:
        cookies = pickle.load(cookiesfile)
        for cookie in cookies:
            driver.add_cookie(cookie)

# 读取yaml
def get_yaml(filename):
    # 文件
    file = filename
    with open(file, 'r', encoding='utf-8') as f:
        # 读取yaml文件
        data = yaml.load(f, Loader=yaml.FullLoader)
        # print(data)
    return data

# 读取Excel
def get_data(filename, Sheet):
    # 文件
    file = filename
    wb = openpyxl.load_workbook(file)
    #选择哪个表格
    sheet = wb[Sheet]
    #
    # print(sheet)
    my_data = []
    # 遍历excel表数据
    for row in sheet.iter_rows(values_only=True):
        # print(row)
        my_data.append(row)  # 如果你想把整行的数据作为一个列表添加到my_data中
    print(my_data)

    return my_data
    # 如果只是打印数据，则不需要my_data列表
    # for row in sheet.iter_rows(values_only=True):
    #     for value in row:
    #         print(value)
    #         my_data.append(value)
    # print(my_data)
# 读取json
def get_json_data(filename):
    with open(filename) as j:
        # 将json读取的数据放进列表中
        my_data = []

        load = json.load(j)
        print(load)
        print(load['keys'])
        my_data.extend(load['keys'])
    return my_data



# def path():
#     # 项目名称
#     pro_name = 'jike_Project'
#     # 本文件路径
#     dirname = os.path.dirname(__file__)
#     # 项目名称的下标
#     pro_name_index = dirname.index(pro_name)
#     # print(dirname)
#     # print(pro_name_index)
#     # 截取项目名称前的路径(换电脑了，可以获取新的路径进行拼接) 再加上 项目名 + 要存放的文件夹
#     path_os = dirname[ : pro_name_index + len(pro_name)] + os.sep + 'logs'
#     print(path_os)
# if __name__ == '__main__':
#     path()